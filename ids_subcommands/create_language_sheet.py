"""
Generates etc/*.xlsx file for editing of passed language ID
"""
import pathlib
import openpyxl
import pycldf
import pylexibank

from cldfbench.cli_util import with_dataset, add_dataset_spec


def register(parser):
    add_dataset_spec(parser)
    parser.add_argument(
        '--lang-id',
        type=int,
        help='CDLF Language_ID',
        default=None,
    )
    parser.add_argument(
        '--add-representations',
        type=str,
        help='Comma-separated list of additional representations',
        default='',
        required=False,
    )


def run(args):
    with_dataset(args, create_language_sheet)


def create_language_sheet(dataset, args):
    assert args.lang_id

    ds = dataset.cldf_reader()

    desired_lang_id = str(args.lang_id)
    desired_fields = ['Parameter_ID', 'Form', 'Comment', 'AlternativeValues']

    lg = ds.get_row('LanguageTable', desired_lang_id)
    assert lg
    args.log.info('creating Excel sheet for "{}" ...'. format(lg['Name']))

    params_id2name = dict((p['ID'], p['Name']) for p in ds['ParameterTable'])

    lg_path_name = lg['Name'].replace(' ', '_').replace('(', '').replace(')', '')
    xlsx_path = dataset.etc_dir / '{}-{}.xlsx'.format(lg_path_name, desired_lang_id)

    nr_additional_cols = 0
    header = ['Chapter', 'Entry', 'English'] + lg['Representations']
    if args.add_representations:
        colnames = list(filter(None, map(str.strip, args.add_representations.split(','))))
        header += colnames
        nr_additional_cols = len(colnames)
    header += ['Comment']
    header_letters = [openpyxl.utils.get_column_letter(i+1) for i in range(len(header))]

    wb = openpyxl.Workbook()
    wa = wb.active
    wa.append(header)
    header_font = openpyxl.styles.Font(color='0000FF', bold=True, size=14)
    cell_font = openpyxl.styles.Font(size=14)
    wa.freeze_panes = "A2"
    for c in header_letters:
        wa["{}1".format(c)].font = header_font

    row_cnt = 1
    for row in pylexibank.progressbar(ds['FormTable'], desc='Extracting forms'):
        if row['Language_ID'] != desired_lang_id:
            continue
        row_cnt += 1
        d = list(map(int, row['Parameter_ID'].split('-')))
        d.append(params_id2name[row['Parameter_ID']])
        d.append(row['Form'])
        for v in row['AlternativeValues']:
            d.append(v)
        if nr_additional_cols:
            d += [''] * nr_additional_cols
        d.append(row['Comment'])
        wa.append(d)
        for c in header_letters:
            wa["{}{}".format(c, row_cnt)].font = cell_font

    for i in range(2, len(header)):
        wa.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = 36

    wb.save(xlsx_path)
